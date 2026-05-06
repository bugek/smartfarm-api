"""Recover versioned GAP baseline artifacts for OME-107 / OME-110.

The live shared compliance workspace was advanced in place from the published
1.0.0 baseline to later 1.1.0 and 1.2.0 revisions. This script recreates a
versioned recovery package for the 1.0.0 baseline using the authoritative
release trail recorded in OME-6 / OME-21 / OME-22 / OME-35 / OME-107 / OME-108.

Outputs land under:
  compliance/recovered/ome107-baseline-1.0.0/
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
COMPLIANCE = ROOT / "compliance"
OUT = COMPLIANCE / "recovered" / "ome107-baseline-1.0.0"
XLSX = COMPLIANCE / "_ams-checklist-v6.2.xlsx"
LIVE_RULESET = COMPLIANCE / "usda-hgap-v1.json"
LIVE_OVERLAY = COMPLIANCE / "usda-hgap-overlay-leafy-greens-v1.json"

OUT_REPORT = OUT / "ams-coverage-report.md"
OUT_SECTION_MD = OUT / "_ams-coverage-section-b.md"
OUT_SECTION_JSON = OUT / "_ams-coverage-section-b.json"
OUT_MANIFEST = OUT / "baseline-manifest.json"
OUT_README = OUT / "README.md"


# Exact later-release row boundaries recovered from OME-107 + OME-108.
PHASE_4_1_ROWS = {
    "G-1.1",
    "G-1.1.a",
    "G-1.2.a",
    "G-1.3",
    "G-2.3",
    "G-2.4",
    "G-2.4.a",
    "G-3.3.a",
    "G-5.1",
    "G-5.2",
    "G-5.3",
    "G-6.1.c",
    "G-6.2",
    "G-8.1",
    "G-8.1.a",
    "G-8.1.b",
    "G-8.2",
    "G-10.5",
    "G-10.9",
    "G-10.10",
    "G-10.11",
    "G-10.12",
    "G-10.13",
    "G-10.14",
    "G-10.15",
    "G-10.17",
    "G-10.18",
    "G-11.4",
    "G-11.5",
    "G-11.6",
    "G-11.7",
    "G-11.8",
    "G-12.1",
    "G-12.2",
    "G-13.1",
    "G-13.2",
    "G-13.2.a",
    "G-13.2.b",
    "G-14.1.a",
    "G-14.1.b",
    "F-11.1",
    "F-11.2",
}

PHASE_4_2_ROWS = {
    "F-1.2",
    "F-1.3",
    "F-4.4",
    "F-4.5",
    "F-6.3",
    "F-7.2.c",
    "F-7.3",
    "F-7.5",
    "F-7.6",
    "F-9.4",
    "F-9.6",
    "F-10.1",
    "F-10.2",
    "F-10.3",
    "F-10.4",
    "F-11.4",
    "F-12.2",
}

assert len(PHASE_4_1_ROWS) == 42
assert len(PHASE_4_2_ROWS) == 17
assert PHASE_4_1_ROWS.isdisjoint(PHASE_4_2_ROWS)


def normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_ams_id(req: str, prefix: str) -> bool:
    if not req.startswith(prefix + "-"):
        return False
    return "." in req[len(prefix) + 1 :]


def load_live_mappings() -> dict[str, list[str]]:
    data = json.loads(LIVE_RULESET.read_text(encoding="utf-8-sig"))
    mapping: dict[str, list[str]] = {}
    for ctl in data["controls"]:
        for ams_id in ctl.get("ams_ids", []) or []:
            mapping.setdefault(ams_id, []).append(ctl["id"])
    return mapping


def classify_row(ams_id: str) -> tuple[str, str]:
    if ams_id in PHASE_4_1_ROWS:
        return ("unmapped_at_1_0_0", "1.1.0")
    if ams_id in PHASE_4_2_ROWS:
        return ("unmapped_at_1_0_0", "1.2.0")
    return ("mapped_at_1_0_0", "1.0.0")


def emit_json(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def count_where(rows: Iterable[dict], value: str) -> int:
    return sum(1 for row in rows if row["status"] == value)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)

    live_ruleset = json.loads(LIVE_RULESET.read_text(encoding="utf-8-sig"))
    live_overlay = json.loads(LIVE_OVERLAY.read_text(encoding="utf-8-sig"))
    live_mappings = load_live_mappings()

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    worksheets = [
        ("Checklist-General Questions", "General Questions", "G"),
        ("Checklist-Field Ops", "Field Operations", "F"),
    ]

    section_rows: list[dict] = []
    summary = {
        "General Questions": {"total": 0, "mapped_at_1_0_0": 0, "introduced_in_1_1_0": 0, "introduced_in_1_2_0": 0},
        "Field Operations": {"total": 0, "mapped_at_1_0_0": 0, "introduced_in_1_1_0": 0, "introduced_in_1_2_0": 0},
    }

    for sheet_name, section_name, prefix in worksheets:
        ws = wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            req = normalize(ws.cell(row=r, column=1).value)
            text = normalize(ws.cell(row=r, column=2).value)
            if not is_ams_id(req, prefix):
                continue

            status, introduced_in = classify_row(req)
            mapped_now_to = live_mappings.get(req, [])
            row = {
                "worksheet": sheet_name,
                "section": section_name,
                "row_number": r,
                "ams_id": req,
                "status": status,
                "introduced_in_revision": introduced_in,
                "mapped_now_to_1_2_0_controls": mapped_now_to,
                "text": text,
            }
            section_rows.append(row)

            bucket = summary[section_name]
            bucket["total"] += 1
            if introduced_in == "1.0.0":
                bucket["mapped_at_1_0_0"] += 1
            elif introduced_in == "1.1.0":
                bucket["introduced_in_1_1_0"] += 1
            else:
                bucket["introduced_in_1_2_0"] += 1

    total_mapped = count_where(section_rows, "mapped_at_1_0_0")
    total_unmapped = len(section_rows) - total_mapped

    manifest = {
        "recovered_for_issue": "OME-110",
        "unblocks_issue": "OME-107",
        "canonical_workspace_root": str(ROOT),
        "canonical_live_artifacts": {
            "ruleset": "compliance/usda-hgap-v1.json",
            "overlay": "compliance/usda-hgap-overlay-leafy-greens-v1.json",
            "coverage_report": "compliance/ams-coverage-report.md",
            "coverage_inventory_md": "compliance/_ams-coverage-section-b.md",
            "coverage_inventory_json": "compliance/_ams-coverage-section-b.json",
            "parser": "scripts/parse_ams_checklist.py",
        },
        "published_1_0_0_identity": {
            "scheme": "USDA_HGAP",
            "version": "3.1",
            "ruleset_revision": "1.0.0",
            "status": "published",
            "overlay_revision": live_overlay.get("ruleset_revision"),
            "overlay_status": live_overlay.get("status"),
            "overlay_applies_to_versions": live_overlay.get("applies_to_versions"),
            "published_from_issue_evidence": [
                "OME-22 conditional approve + close-out",
                "OME-35 mechanical publish fix",
                "OME-69 README 1.0.0 update",
            ],
        },
        "baseline_counts": {
            "controls_total": 46,
            "controls_general_questions": 22,
            "controls_field_operations": 24,
            "ams_rows_mapped_total": total_mapped,
            "ams_rows_mapped_general_questions": summary["General Questions"]["mapped_at_1_0_0"],
            "ams_rows_mapped_field_operations": summary["Field Operations"]["mapped_at_1_0_0"],
            "ams_rows_unmapped_total": total_unmapped,
        },
        "known_0_2_0_control_additions_before_publish": {
            "general_questions": [
                "HGAP.GQ.2.4",
                "HGAP.GQ.2.5",
                "HGAP.GQ.4.3",
                "HGAP.GQ.6.2",
                "HGAP.GQ.10.1",
                "HGAP.GQ.10.2",
            ],
            "field_operations": [
                "HGAP.FO.2.4",
                "HGAP.FO.4.3",
                "HGAP.FO.4.4",
                "HGAP.FO.5.2",
                "HGAP.FO.6.4",
                "HGAP.FO.6.5",
            ],
        },
        "rows_introduced_after_1_0_0": {
            "phase_4_1_revision": "1.1.0",
            "phase_4_1_rows": sorted(PHASE_4_1_ROWS),
            "phase_4_2_revision": "1.2.0",
            "phase_4_2_rows": sorted(PHASE_4_2_ROWS),
        },
        "assumptions": [
            "The exact 1.0.0 full JSON body was superseded in place in the shared workspace and is not present as a separate file on disk.",
            "The 1.0.0 baseline row boundary is recovered from the approved Phase 4.1 and Phase 4.2 issue scopes plus their completion comments.",
            "Rows outside the approved 1.1.0 and 1.2.0 scope lists were already mapped in the published 1.0.0 baseline.",
        ],
        "caveats": [
            "Use this package as the authoritative 1.0.0 baseline for OME-107 release planning and coverage accounting.",
            "Use the live 1.2.0 ruleset only as a reference for wording patterns and current traceability, not as the 1.0.0 file-of-record for OME-107.",
        ],
    }
    emit_json(OUT_MANIFEST, manifest)

    lines = [
        "# OME-107 Baseline Recovery Package",
        "",
        "This folder restores the authoritative **1.0.0 baseline context** needed by",
        "`OME-107` after the shared compliance workspace was advanced in place to later",
        "`1.1.0` and `1.2.0` revisions.",
        "",
        "Canonical workspace",
        f"- `{ROOT}`",
        "",
        "Live files of record",
        "- `compliance/usda-hgap-v1.json` (current live base ruleset; now `1.2.0`)",
        "- `compliance/usda-hgap-overlay-leafy-greens-v1.json` (published overlay `0.1.0`)",
        "- `compliance/ams-coverage-report.md`",
        "- `compliance/_ams-coverage-section-b.{md,json}`",
        "- `scripts/parse_ams_checklist.py`",
        "",
        "Recovered baseline artifacts in this folder",
        "- `baseline-manifest.json` - published 1.0.0 identity, counts, provenance, and exact later-release row boundaries.",
        "- `ams-coverage-report.md` - 1.0.0 baseline coverage summary for OME-107 planning.",
        "- `_ams-coverage-section-b.{md,json}` - full AMS row inventory classified as `mapped_at_1_0_0` vs introduced later in `1.1.0` / `1.2.0`.",
        "",
        "Important caveat",
        "- The exact standalone `1.0.0` JSON body is not preserved on disk after the in-place promotion to `1.2.0`.",
        "- This package therefore recovers the authoritative **baseline release boundary** from the issue trail instead of fabricating a synthetic full JSON file.",
        "- For `OME-107`, the critical source of truth is which AMS rows were still unmapped at `1.0.0`; those rows are captured exactly here.",
        "",
    ]
    OUT_README.write_text("\n".join(lines) + "\n", encoding="utf-8")

    md_lines = [
        "# AMS Combined Checklist v6.2 - Recovered 1.0.0 Baseline Coverage",
        "",
        "**Recovered for:** `OME-110` to unblock `OME-107`.",
        "**Canonical workspace:** `_default/` shared SmartFarm project workspace.",
        "**Baseline release identity:** base `usda-hgap@3.1` rev `1.0.0`, overlay `usda-hgap-leafy-greens@3.1` rev `0.1.0`, both published.",
        "**Recovery method:** derived from the authoritative issue trail because the live `compliance/usda-hgap-v1.json` file was advanced in place to `1.2.0`.",
        "",
        "## Headline numbers",
        "",
        f"- **126** total AMS base rows across General Questions + Field Operations.",
        f"- **{total_mapped}** rows mapped in the published `1.0.0` baseline.",
        f"- **{len(PHASE_4_1_ROWS)}** rows were added in `1.1.0` (the exact Phase 4.1 target for `OME-107`).",
        f"- **{len(PHASE_4_2_ROWS)}** rows remained deferred until `1.2.0`.",
        f"- **{total_unmapped}** total rows were therefore unmapped at `1.0.0`.",
        "",
        "## Section summary",
        "",
        f"- General Questions mapped at `1.0.0`: **{summary['General Questions']['mapped_at_1_0_0']} / {summary['General Questions']['total']}**.",
        f"- Field Operations mapped at `1.0.0`: **{summary['Field Operations']['mapped_at_1_0_0']} / {summary['Field Operations']['total']}**.",
        "",
        "## Exact rows that `OME-107` must treat as post-1.0.0 Phase 4.1 scope",
        "",
        ", ".join(sorted(PHASE_4_1_ROWS)),
        "",
        "## Exact rows that stayed deferred until 1.2.0 / Phase 4.2",
        "",
        ", ".join(sorted(PHASE_4_2_ROWS)),
        "",
        "## Notes",
        "",
        "- Rows classified as `mapped_at_1_0_0` were already covered before the approved Phase 4 expansion work began.",
        "- Rows classified as `introduced_in_revision = 1.1.0` or `1.2.0` come directly from the approved issue scopes and completion comments for `OME-107` and `OME-108`.",
        "- Use the versioned inventory in `_ams-coverage-section-b.{md,json}` for row-by-row implementation planning.",
        "",
    ]
    OUT_REPORT.write_text("\n".join(md_lines) + "\n", encoding="utf-8")

    emit_json(OUT_SECTION_JSON, {"rows": section_rows})

    table_lines = [
        "# AMS Combined Checklist v6.2 - Recovered 1.0.0 Inventory",
        "",
        "Each row is classified against the published `1.0.0` baseline, not the current live `1.2.0` file.",
        "",
    ]
    for section_name in ("General Questions", "Field Operations"):
        rows = [row for row in section_rows if row["section"] == section_name]
        table_lines.extend(
            [
                f"## {section_name}",
                "",
                "| Worksheet Row | AMS ID | 1.0.0 Status | Introduced In | Live 1.2.0 Mapping | Text |",
                "| --- | --- | --- | --- | --- | --- |",
            ]
        )
        for row in rows:
            mapped_to = ", ".join(row["mapped_now_to_1_2_0_controls"]) or "-",
            mapped_to_text = mapped_to[0]
            text = row["text"].replace("|", "\\|")
            table_lines.append(
                f"| {row['row_number']} | {row['ams_id']} | {row['status']} | "
                f"{row['introduced_in_revision']} | {mapped_to_text} | {text} |"
            )
        table_lines.append("")

    OUT_SECTION_MD.write_text("\n".join(table_lines), encoding="utf-8")

    print(f"Recovered OME-107 baseline package in {OUT}")


if __name__ == "__main__":
    main()
